#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  editor.py
#
#  Copyright 2020 iwoithe <iwoithe@just42.net>
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#
#

import os
import csv
import sys
import openpyxl

from openpyxl.styles.fonts import Font

from PyQt5 import QtCore, QtGui, QtWidgets

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (QApplication, QWidget, QDockWidget,
                             QHBoxLayout, QVBoxLayout, QTabWidget,
                             QPlainTextEdit)

class Editor(QDockWidget):

    files_open = []

    def __init__(self):
        super().__init__("Editor")

        self.setup_ui()

    def setup_ui(self):
        self.tabs = QTabWidget()
        self.tabs.setMovable(True)
        self.tabs.setTabsClosable(True)
        self.tabs.tabCloseRequested.connect(self.close_tab)

        # Load Files
        if (len(self.files_open) > 0):
            self.load_files(self.files_open)
        else:
            self.new_file()

        # Main Layout
        widget = QWidget()
        layout = QVBoxLayout()

        layout.addWidget(self.tabs)

        widget.setLayout(layout)
        self.setWidget(widget)

    def close_tab(self, i):
        self.tabs.removeTab(i)

    def new_file(self, names=None):
        if (names == None):
            file_tab = QWidget()
            layout = QHBoxLayout()

            file_tab.editor = QPlainTextEdit()

            layout.addWidget(file_tab.editor)
            file_tab.setLayout(layout)
            self.tabs.addTab(file_tab, "untitled.rgg")

            self.tabs.setCurrentWidget(file_tab)

        else:
            file_tab = QWidget()
            layout = QHBoxLayout()

            file_tab.editor = QPlainTextEdit()

            for name in names:
                file_tab.editor.insertPlainText(name)

            layout.addWidget(file_tab.editor)
            file_tab.setLayout(layout)
            self.tabs.addTab(file_tab, "imported*")

            self.tabs.setCurrentWidget(file_tab)

    def save_file(self, file):
        current_file = self.tabs.currentWidget()
        text = current_file.editor.toPlainText()

        with open(file, 'w') as f:
            f.write(text)

        self.tabs.setTabText(self.tabs.indexOf(current_file), os.path.basename(file))

    def load_files(self, files):
        for file in files:
            file_tab = QWidget()
            layout = QHBoxLayout()

            file_tab.editor = QPlainTextEdit()

            with open(file) as f:
                for line in f:
                    file_tab.editor.insertPlainText(line)

            layout.addWidget(file_tab.editor)

            file_tab.setLayout(layout)
            self.tabs.addTab(file_tab, os.path.basename(file))

            self.tabs.setCurrentWidget(file_tab)

    def import_files(self, files, file_type):
        names = []
        for file in files:
            if file_type == "Comma Seperated Values (*.csv)":
                with open(file) as f:
                    reader = csv.reader(f, delimiter=",")
                    for row in reader:
                        name = row[0] + "\n"
                        names.append(name)

                self.new_file(names)

            elif file_type == "Microsoft Excel Spreadsheets (*.xlsx)":
                wb = openpyxl.load_workbook(file)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    num_names = sheet.max_row
                    for row in range(1, num_names + 1):
                        name = sheet.cell(row, 1).value + "\n"
                        names.append(name)

                    self.new_file(names)
                    names = []

            else:
                pass

    def export_file(self, groups, file, file_type):
        if file_type == "Comma Seperated Values (*.csv)":
            with open(file, 'w', newline='') as f:
                writer = csv.writer(f, delimiter=',')
                for group in groups:
                    people = ""
                    for person in group[1]:
                        people += person + ", "

                    row = str(group[0]) + ", " + people
                    row = row.split(",")
                    writer.writerow(row)

        elif file_type == "Microsoft Excel Spreadsheets (*.xlsx)":
            # Setup the workbook and sheet
            wb = openpyxl.Workbook()
            sheet = wb['Sheet']
            sheet.title = "Generated Groups"
            # Setup the header
            # TODO: Centre the headers
            header_font = Font(bold=True)

            people_per_group = len(groups[0][1]) + 1
            sheet.cell(1, 1).font = header_font
            sheet.cell(1, 1).value = "#"
            sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=people_per_group)
            sheet.cell(1, 2).font = header_font
            sheet.cell(1, 2).value = "Groups"
            # Put the data in the cells below the header
            for group in groups:
                people = ""
                for person in group[1]:
                    people += person + ", "

                row = str(group[0]) + ", " + people
                row = row.split(",")
                row[0] = int(row[0])
                for column in row:
                    sheet.cell(group[0] + 1, row.index(column) + 1).value = column

            wb.save(file)

        else:
            pass


if __name__ == '__main__':
    app = QApplication(sys.argv)
    editor = Editor()
    editor.show()
    sys.exit(app.exec_())
